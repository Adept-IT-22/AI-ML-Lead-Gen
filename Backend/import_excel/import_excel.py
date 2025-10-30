import logging
from typing import List
from openpyxl import load_workbook

logger = logging.getLogger()
logging.basicConfig(level=logging.INFO)

def main(file: str)->List[str]:
    print("Starting file input...")
    
    #Load file
    workbook = load_workbook(file)

    #Get first worksheet (the one on the far left in the file)
    list_of_worksheets = workbook.worksheets
    desired_sheet = list_of_worksheets[0]

    #Open and read the first row, goal is to get the index website column
    x = {}
    for i, row in enumerate(desired_sheet.iter_rows(values_only=True)):
        x[i] = row
    
    website_col_index = None
    headers = list(x[0])
    for i, val in enumerate(headers):
        if val != None and 'website' in val.lower():
            website_col_index = i + 1 #Because the iter_cols method is 1-indexed

    if website_col_index == None:
        logger.error('No website url column found')
        return []

    website_urls = []
    for row in desired_sheet.iter_cols(min_col=website_col_index, max_col=website_col_index):
        for cell in row:
            website_urls.append(cell.value)

    return website_urls

if __name__ == "__main__":
    main()    