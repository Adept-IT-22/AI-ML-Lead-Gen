import os
import copy
import httpx
import asyncio
import asyncpg
import logging
from typing import List
from dotenv import load_dotenv
from utils.data_structures.news_data_structure import fetched_funding_data
from openpyxl import load_workbook
from orchestration.enrichment import main as enrichment_main
from orchestration.storage import main as storage_main
from orchestration.scoring import main as scoring_main
from orchestration.outreach import main as outreach_main
from enrichment_module.organization_search import org_search

#Set up DB_URL
load_dotenv(override=True)
DB_URL = os.getenv("MOCK_DATABASE_URL")

#Set up logger
logger = logging.getLogger()
logging.basicConfig(level=logging.INFO)

#Expected file names
files_to_import = {
    'FUNDED_IQ':'fundediq',
    'ADEPT_WEBSITE': 'apollo-accounts',
    'GERMANY': 'germany' # Hannover Messe
    }

#Column names
column_names = {
    'FUNDED_IQ': ['Website', 'company', 'funding type', 'funding amount'],
    'ADEPT_WEBSITE': ['website', 'company name', 'latest funding', 'latest funding amount'],
    'GERMANY': ['company name', 'location', 'founded', 'city', 'website']
    }

# ===============GET FILE NAME=============
def get_file_name(file)->str:
    logger.info("Getting file name")

    filename= file.filename
    return filename

# ===============EXTRACT DESIRED SHEET=============

def get_desired_sheet(file)->List[str]:
    logger.info("Getting newest sheet")
    
    #Load file
    workbook = load_workbook(file)

    #Get first worksheet (the one on the far left in the file)
    list_of_worksheets = workbook.worksheets
    desired_sheet = list_of_worksheets[0]

    return desired_sheet

# ===============EXTRACT COLUMN DATA=============

def extract_column_data(part_filename: str, actual_filename, desired_sheet, column_name:str)->List[str]:
    column_data = []

    if part_filename in actual_filename.lower():
        column_data = return_column_data(desired_sheet, column_name)

    return column_data

def return_column_data(desired_sheet, column: str)->List[str]:
    #Put rows in this dictionary with numbers as keys
    row_dict = {}
    for i, row in enumerate(desired_sheet.iter_rows(values_only=True)):
        row_dict[i] = row
    
    column_index = 0
    headers = list(row_dict[0])

    for i, val in enumerate(headers):
        if val != None and column.lower() in val.lower():
            column_index = i + 1 #Because the iter_cols method is 1-indexed

    if column_index == 0:
        logger.error('No such column exists')
        return []

    column_data = []
    for row in desired_sheet.iter_cols(min_col=column_index, max_col=column_index):
        for cell in row:
            column_data.append(cell.value)

    #Remove the 1st entry as it's row 1 in the excel sheet which is the column title
    final_column_data = column_data[1:]

    return final_column_data


# =========CONVERT DATA INTO NORMALIZED DATA========
def create_normalized_data(file):
    filename = get_file_name(file)
    desired_sheet = get_desired_sheet(file)

    funding_data_structure = copy.deepcopy(fetched_funding_data)

    # Example for FUNDED_IQ file
    if 'fundediq' in filename.lower():
        company_names = extract_column_data('fundediq', filename, desired_sheet, 'company')
        funding_rounds = extract_column_data('fundediq', filename, desired_sheet, 'funding type')
        amounts = extract_column_data('fundediq', filename, desired_sheet, 'funding amount')

        funding_data_structure['company_name'].extend(company_names)
        funding_data_structure['funding_round'].extend(funding_rounds)
        funding_data_structure['amount_raised'].extend(amounts)
        funding_data_structure['source'].append('Funded IQ')

    # Example for ADEPT_WEBSITE file
    if 'apollo-accounts' in filename.lower():
        company_names = extract_column_data('apollo-accounts', filename, desired_sheet, 'company name')
        funding_rounds = extract_column_data('apollo-accounts', filename, desired_sheet, 'latest funding')
        amounts = extract_column_data('apollo-accounts', filename, desired_sheet, 'latest funding amount')

        funding_data_structure['company_name'].extend(company_names)
        funding_data_structure['funding_round'].extend(funding_rounds)
        funding_data_structure['amount_raised'].extend(amounts)
        funding_data_structure['source'].append('Adept Website')

    # Example for GERMANY file
    if files_to_import.get('GERMANY', '') in filename.lower():
        company_names = extract_column_data('germany', filename, desired_sheet, 'company name')
        countries = extract_column_data('germany', filename, desired_sheet, 'location')
        cities = extract_column_data('germany', filename, desired_sheet, 'city')
    

        funding_data_structure['company_name'].extend(company_names)
        funding_data_structure['country'].extend(countries)
        funding_data_structure['city'].extend(cities)
        funding_data_structure['source'].append('Hannover Messe 2026 Lead List')

    logger.info([funding_data_structure])
    return [funding_data_structure]

# ========ENRICH, SCORE, STORE, OUTREACH==========

async def main(file):
    ORGANIZATION_SEARCH_URL = "https://api.apollo.io/api/v1/organizations/search"
    from config.apollo_config import headers as APOLLO_HEADERS
    import aiofiles
    import json

    normalized_data = create_normalized_data(file)
    #async with httpx.AsyncClient(timeout=10.0) as client:
        #tasks = [org_search(client, comp, api_url = ORGANIZATION_SEARCH_URL, headers = APOLLO_HEADERS) for comp in normalized_data[0].get('company_name', [])]
        #results = await asyncio.gather(*tasks)

    #logger.info("RESULTS: %r", results)

    # ===========QUEUE CREATION ===============
    normalization_to_enrichment= asyncio.Queue()
    normalization_to_storage= asyncio.Queue()
    enrichment_to_storage_queue = asyncio.Queue()

    await normalization_to_enrichment.put(normalized_data)
    await normalization_to_storage.put(normalized_data)

    async with asyncpg.create_pool(dsn=DB_URL, min_size=1, max_size=100) as pool:

        enrichment_module_queue = await enrichment_main(
            normalization_to_enrichment_queue=normalization_to_enrichment,
            enrichment_to_storage_queue=enrichment_to_storage_queue
        )
        
        #enrichment_data = await enrichment_module_queue.get()

        #async with aiofiles.open("xxx.txt", "w") as file:
            #await file.write(json.dumps(enrichment_data, indent=2))

        await storage_main(
            pool,
            normalization_to_storage_queue=normalization_to_storage,
            enrichment_to_storage_queue=enrichment_module_queue
        )

        #await scoring_main(
            #pool
        #)

        #await outreach_main(
            #pool
        #)

if __name__ == "__main__":
    asyncio.run(main())